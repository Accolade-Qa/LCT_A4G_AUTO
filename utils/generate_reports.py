import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ================= CONFIG =================
ROOT = Path.cwd()

_reports_upper = ROOT / "Reports"
_reports_lower = ROOT / "reports"

REPORT_DIR = _reports_upper if _reports_upper.exists() else _reports_lower
REPORT_DIR.mkdir(parents=True, exist_ok=True)

_json_upper = _reports_upper / "json" / "report.json"
_json_lower = _reports_lower / "report.json"
_manual_upper = _reports_upper / "test_report.xlsx"
_manual_lower = _reports_lower / "test_report.xlsx"

JSON_PATH = None
for path in [_json_upper, _json_lower]:
    if path.exists():
        JSON_PATH = path
        break

if JSON_PATH is None:
    JSON_PATH = REPORT_DIR / "report.json"

HTML_PATH = REPORT_DIR / "report.html"
EXCEL_PATH = REPORT_DIR / "report.xlsx"
MANUAL_EXCEL_PATH = None
for path in [_manual_upper, _manual_lower]:
    if path.exists():
        MANUAL_EXCEL_PATH = path
        break

if MANUAL_EXCEL_PATH is None:
    MANUAL_EXCEL_PATH = REPORT_DIR / "test_report.xlsx"


# ================= HTML TEMPLATE =================
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Automation Test Report</title>
<style>
body {
    font-family: Segoe UI, Arial, sans-serif;
    background: #f8fafc;
    color: #0f172a;
    margin: 0;
    padding: 24px;
}
h1 { margin-bottom: 4px; }
.meta { color: #475569; margin-top: 0; }
.cards {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 15px;
    margin: 24px 0;
}
.card {
    padding: 16px;
    border-radius: 8px;
    background: white;
    border: 1px solid #e2e8f0;
    box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06);
}
.pass { border-left: 5px solid #22c55e; }
.fail { border-left: 5px solid #ef4444; }
.skip { border-left: 5px solid #f59e0b; }
.value { font-size: 28px; font-weight: 700; margin-top: 6px; }
table {
    width: 100%;
    border-collapse: collapse;
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    overflow: hidden;
    table-layout: fixed;
}
th {
    background: #e2e8f0;
    color: #0f172a;
    text-align: left;
    font-size: 13px;
}
td, th {
    padding: 10px 12px;
    border-bottom: 1px solid #e2e8f0;
    vertical-align: top;
    word-wrap: break-word;
}
.status-pass { color: #166534; font-weight: 700; }
.status-fail { color: #991b1b; font-weight: 700; }
.status-skipped { color: #92400e; font-weight: 700; }
.muted { color: #64748b; }
</style>
</head>
<body>
<h1>Automation Test Report</h1>
<p class="meta">Generated at {{ time }}</p>

<div class="cards">
<div class="card"><div>Total</div><div class="value">{{ total }}</div></div>
<div class="card pass"><div>Passed</div><div class="value">{{ passed }}</div></div>
<div class="card fail"><div>Failed</div><div class="value">{{ failed }}</div></div>
<div class="card skip"><div>Skipped</div><div class="value">{{ skipped }}</div></div>
</div>

<table>
<tr>
    <th style="width: 22%;">Test Case Name</th>
    <th style="width: 26%;">Expected</th>
    <th style="width: 26%;">Actual</th>
    <th style="width: 9%;">Result</th>
    <th style="width: 8%;">Duration</th>
    <th style="width: 9%;">Message</th>
</tr>
{% for t in tests %}
<tr>
    <td>{{ t.name }}</td>
    <td>{{ t.expected }}</td>
    <td>{{ t.actual }}</td>
    <td class="status-{{ t.status }}">{{ t.status|upper }}</td>
    <td>{{ t.duration }}{% if t.duration != "" %}s{% endif %}</td>
    <td class="muted">{{ t.message }}</td>
</tr>
{% endfor %}
</table>
</body>
</html>
"""


def _clean(value, limit=500):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).replace("\r", " ").replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value[:limit]


def _test_case_name(nodeid):
    if not nodeid:
        return ""
    return nodeid.split("::")[-1]


def _normalize_status(value):
    value = str(value or "").strip().lower()
    if value in {"passed", "pass"}:
        return "pass"
    if value in {"failed", "fail"}:
        return "fail"
    return "skipped"


def _duration_seconds(test_result):
    duration = test_result.get("duration")
    if duration is not None:
        return round(duration, 2)

    total = 0
    found = False
    for section in ("setup", "call", "teardown"):
        section_data = test_result.get(section) or {}
        section_duration = section_data.get("duration")
        if section_duration is not None:
            total += section_duration
            found = True

    return round(total, 2) if found else ""


def _property_map(test_result):
    properties = {}
    raw_properties = test_result.get("user_properties") or test_result.get("properties") or []

    if isinstance(raw_properties, dict):
        return {str(key).lower(): value for key, value in raw_properties.items()}

    for item in raw_properties:
        if isinstance(item, dict):
            name = item.get("name") or item.get("key")
            value = item.get("value")
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            name, value = item[0], item[1]
        else:
            continue

        if name:
            properties[str(name).lower()] = value

    return properties


def _load_manual_results():
    """Read richer expected/actual rows written by utils.excel_report.write_result."""
    if not MANUAL_EXCEL_PATH.exists():
        return {}

    try:
        df = pd.read_excel(MANUAL_EXCEL_PATH)
    except Exception as exc:
        print(f"Warning: Could not read manual Excel report {MANUAL_EXCEL_PATH}: {exc}")
        return {}

    required = {"Test Name", "Expected", "Actual", "Status"}
    missing = required - set(df.columns)
    if missing:
        print(f"Warning: Manual report missing columns: {missing}")
        return {}

    results = {}
    for _, row in df.iterrows():
        test_name = _clean(row.get("Test Name"))
        if not test_name:
            continue
        results[test_name] = {
            "expected": _clean(row.get("Expected")),
            "actual": _clean(row.get("Actual")),
            "status": _normalize_status(row.get("Status")),
            "message": _clean(row.get("Error"), 300),
        }
    return results


def _extract_expected_actual(longrepr, outcome):
    """Best-effort extraction for tests that do not call write_result()."""
    text = _clean(longrepr, 2000)

    if not text:
        return "Not recorded by test", "Not recorded by test", ""

    patterns = [
        r"Expected:\s*(?P<expected>.*?)[,;]\s*(?:Got|Actual):\s*(?P<actual>.*)",
        r"Expected\s+'(?P<expected>.*?)'\s*,?\s*(?:but\s+)?got\s+'(?P<actual>.*?)'",
        r"Expected\s+(?P<expected>.*?)\s*,\s*got\s+(?P<actual>.*)",
        r"expected\s+(?P<expected>.*?)\s+but\s+got\s+(?P<actual>.*)",
        r"Expected\s+URL\s+to\s+be\s+'(?P<expected>.*?)',\s+got\s+(?P<actual>.*)",
        r"Expected\s+page\s+title\s+'(?P<expected>.*?)',\s+but\s+got\s+'(?P<actual>.*?)'",
        r"Expected\s+title\s+'(?P<expected>.*?)',\s+but\s+got\s+'(?P<actual>.*?)'",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return (
                _clean(match.group("expected"), 500),
                _clean(match.group("actual"), 500),
                text[:300],
            )

    if "AssertionError:" in text:
        message = text.split("AssertionError:", 1)[-1].strip()
        return "Assertion condition should pass", _clean(message, 500), text[:300]

    return "Not recorded by test", _clean(text, 500), text[:300]


def _build_test_rows(data):
    manual_results = _load_manual_results()
    tests = []
    seen = set()
    counts = {"passed": 0, "failed": 0, "skipped": 0}

    for t in data.get("tests", []):
        nodeid = t.get("nodeid", "")
        test_name = _test_case_name(nodeid)
        status = _normalize_status(t.get("outcome", ""))
        properties = _property_map(t)
        
        # 🔹 PRIORITY 1: Check properties (set by report_case fixture or pytest hook)
        properties_expected = _clean(properties.get("expected"))
        properties_actual = _clean(properties.get("actual"))
        properties_message = _clean(properties.get("message"), 300)
        
        # 🔹 PRIORITY 2: Extract from error/longrepr only if properties don't have expected/actual
        if properties_expected or properties_actual:
            # Properties has explicit values - use them
            expected = properties_expected
            actual = properties_actual
            message = properties_message or ""
        else:
            # Fall back to extracting from error message
            expected, actual, message = _extract_expected_actual(
                t.get("longrepr", ""), t.get("outcome", "")
            )
            # Override message with property if available
            message = properties_message or message
        
        # 🔹 PRIORITY 3: Update status from properties if available
        status = _normalize_status(properties.get("result") or properties.get("status") or status)

        # 🔹 PRIORITY 4: Manual results (from write_result) override everything
        manual = manual_results.get(test_name)
        if manual:
            expected = manual["expected"] or expected
            actual = manual["actual"] or actual
            status = manual["status"] or status
            message = manual["message"] or message

        counts[{"pass": "passed", "fail": "failed"}.get(status, "skipped")] += 1
        seen.add(test_name)
        
        # 🔍 Debug logging for "Not recorded by test" issues
        if "Not recorded by test" in (expected or actual):
            print(f"⚠️  DEBUG: Test '{test_name}' has 'Not recorded by test' in expected/actual")
            print(f"   - has properties: {bool(properties)}")
            print(f"   - has manual: {bool(manual)}")
            print(f"   - longrepr length: {len(str(t.get('longrepr', '')))}")
        
        tests.append(
            {
                "name": test_name,
                "nodeid": nodeid,
                "expected": expected,
                "actual": actual,
                "status": status,
                "duration": _duration_seconds(t),
                "message": _clean(message, 300),
            }
        )

    for test_name, manual in manual_results.items():
        if test_name in seen:
            continue
        status = manual["status"]
        counts[{"pass": "passed", "fail": "failed"}.get(status, "skipped")] += 1
        tests.append(
            {
                "name": test_name,
                "nodeid": test_name,
                "expected": manual["expected"],
                "actual": manual["actual"],
                "status": status,
                "duration": "",
                "message": manual["message"],
            }
        )

    return tests, counts, len(tests)


# ================= STEP 1: RUN PYTEST =================
def run_pytest():
    REPORT_DIR.mkdir(exist_ok=True)

    cmd = [
        sys.executable,
        "-m",
        "pytest",
        "--json-report",
        f"--json-report-file={JSON_PATH}",
    ]

    print("Running pytest...")
    result = subprocess.run(cmd, cwd=ROOT)
    return result.returncode


# ================= STEP 2: PROCESS JSON =================
def process_json():
    """Process JSON report and merge manual expected/actual rows when present."""
    data = {"tests": []}

    if not JSON_PATH.exists():
        print(f"Warning: JSON file not found at {JSON_PATH}")
        print(f"Looking in: {_reports_upper / 'json' / 'report.json'}")
        print(f"And: {_reports_lower / 'report.json'}")
        return (*_build_test_rows(data), data)

    try:
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        print(f"JSON loaded from: {JSON_PATH}")
    except json.JSONDecodeError as exc:
        print(f"Warning: Invalid JSON in {JSON_PATH}: {exc}")
        data = {"tests": []}
    except Exception as exc:
        print(f"Warning: Could not read JSON: {exc}")
        data = {"tests": []}

    tests, counts, total = _build_test_rows(data)
    return tests, counts, total, data


# ================= STEP 3: HTML =================
def generate_html(tests, counts, total):
    template = Template(HTML_TEMPLATE)

    html = template.render(
        time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total=total,
        passed=counts["passed"],
        failed=counts["failed"],
        skipped=counts["skipped"],
        tests=tests,
    )

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(html, encoding="utf-8")
    print("HTML generated:", HTML_PATH)


# ================= STEP 4: EXCEL =================
def generate_excel(tests):
    rows = [
        {
            "Test Case Name": t["name"],
            "Expected": t["expected"],
            "Actual": t["actual"],
            "Result": t["status"].upper(),
            "Duration (s)": t["duration"],
            "Message": t["message"],
        }
        for t in tests
    ]

    df = pd.DataFrame(
        rows,
        columns=[
            "Test Case Name",
            "Expected",
            "Actual",
            "Result",
            "Duration (s)",
            "Message",
        ],
    )

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_excel(EXCEL_PATH, index=False, sheet_name="Test Results")
    _style_excel()

    print("Excel generated:", EXCEL_PATH)
    print(f"Total rows: {len(rows)}")


def _style_excel():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Test Results"]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    skip_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    result_col = None
    for cell in ws[1]:
        if cell.value == "Result":
            result_col = cell.column
            break

    if result_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=result_col)
            result = str(cell.value or "").upper()
            if result == "PASS":
                cell.fill = pass_fill
            elif result == "FAIL":
                cell.fill = fail_fill
            else:
                cell.fill = skip_fill
            cell.font = Font(bold=True)

    widths = {
        "A": 36,
        "B": 48,
        "C": 48,
        "D": 12,
        "E": 14,
        "F": 55,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_PATH)


# ================= MAIN =================
def main():
    print("=" * 60)
    print("TEST REPORT GENERATOR")
    print("=" * 60)
    print(f"Working directory: {ROOT}")
    print(f"Report directory: {REPORT_DIR}")
    print(f"JSON path: {JSON_PATH}")
    print(f"Manual expected/actual path: {MANUAL_EXCEL_PATH}")
    print(f"JSON exists: {JSON_PATH.exists()}")
    print("-" * 60)

    run_pytest_flag = os.getenv("RUN_PYTEST", "true").lower() != "false"

    if run_pytest_flag:
        exit_code = run_pytest()
    else:
        print("Skipping pytest (RUN_PYTEST=false)")
        exit_code = 0

    try:
        tests, counts, total, data = process_json()

        print("Generating HTML report...")
        generate_html(tests, counts, total)

        print("Generating Excel report...")
        generate_excel(tests)

        print("\n" + "=" * 60)
        print("ALL REPORTS GENERATED SUCCESSFULLY!")
        print("=" * 60)
        print(f"HTML:  {HTML_PATH} (exists: {HTML_PATH.exists()})")
        print(f"Excel: {EXCEL_PATH} (exists: {EXCEL_PATH.exists()})")
        print(f"JSON:  {JSON_PATH} (exists: {JSON_PATH.exists()})")
        print("=" * 60)

    except Exception as exc:
        print(f"Error processing reports: {exc}")
        try:
            print("Attempting to generate reports with fallback...")
            generate_html([], {"passed": 0, "failed": 0, "skipped": 0}, 0)
            generate_excel([])
        except Exception as fallback_exc:
            print(f"Failed to generate reports: {fallback_exc}")
            return 1

    if exit_code != 0:
        print("Some tests failed")

    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
