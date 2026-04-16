from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# 📁 Create reports folder
folder = "reports"
os.makedirs(folder, exist_ok=True)

# 📄 Common file path (USE THIS EVERYWHERE)
file_path = os.path.join(folder, "test_report.xlsx")


def write_result(test_name, expected, actual, status, error=""):
    # ✅ Create file if not exists
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Test Name", "Time", "Expected", "Actual", "Status", "Error"])
        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    # 🔥 Prevent duplicate BEFORE writing
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == test_name:
            print(f"⚠ Duplicate found for {test_name}, skipping...")
            wb.save(file_path)
            return

    # ✅ Add new row
    ws.append([
        test_name,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        expected,
        actual,
        status,
        error
    ])

    wb.save(file_path)
    print(f"✅ Written: {test_name}")


def remove_duplicates():
    # ✅ Use SAME file path
    if not os.path.exists(file_path):
        print("⚠ File not found, skipping duplicate removal")
        return

    wb = load_workbook(file_path)
    ws = wb.active

    seen = set()
    rows_to_delete = []

    for row in range(2, ws.max_row + 1):
        test_name = ws.cell(row=row, column=1).value

        if test_name in seen:
            rows_to_delete.append(row)
        else:
            seen.add(test_name)

    # 🔥 Delete from bottom
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    wb.save(file_path)
    print("✅ Duplicates removed successfully")